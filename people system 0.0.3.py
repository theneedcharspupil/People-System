from openpyxl import load_workbook
"""
Author: Rain Wang
Date: 2022.1.1
Project name: people system
Verson: 0.0.3
"""
wb = load_workbook('./file/xlsx/class.xlsx')
account_pwd = wb['pwd']
account_account = account_pwd.max_row


def exi():
    print('感谢您的使用，欢迎下次管理')
    exit()


def reversed():
    global people
    input_people = input('请输入人名')
    for row in range(2, people_count + 1):
        if people.cell(row, 1).value == input_people:
            people.delete_rows(row)
            print('删除成功')
    wb.save('./file/xlsx/class.xlsx')


def add_person():
    input_add = input('请输入你要添加的人员姓名')
    phone = input('请输入人员手机号')
    manorwoman = input('请输入人员性别')
    email = input('请输入人员邮箱')
    people.cell(people_count + 1, 1).value = input_add
    people.cell(people_count + 1, 2).value = phone
    people.cell(people_count + 1, 3).value = manorwoman
    people.cell(people_count + 1, 4).value = email
    wb.save('./file/xlsx/class.xlsx')


def cha():
    people_cha = input('请输入查询人员姓名')
    for row in range(1, people_count + 1):
        if people.cell(row, 1).value == people_cha:
            return_cha_name = people.cell(row, 1).value
            return_cha_phone = people.cell(row, 2).value
            return_cha_email = people.cell(row, 4).value
            return_cha_manorwoman = people.cell(row, 3).value
            print('查询人员信息为,姓名{}\t手机号{}\t邮箱{}\t性别{}\t'.format(return_cha_name, return_cha_phone, return_cha_email,
                                                             return_cha_manorwoman))
            inde()
def all_cha():
    for row in range(2, people_count + 1):
        if people.cell(row, 1).value != None:
            return_cha_name2 = people.cell(row, 1).value
            return_cha_phone2 = people.cell(row, 2).value
            return_cha_email2 = people.cell(row, 4).value
            return_cha_manorwoman2 = people.cell(row, 3).value
            print('查询人员信息为,姓名{}\t手机号{}\t邮箱{}\t性别{}\t'.format(return_cha_name2, return_cha_phone2, return_cha_email2,return_cha_manorwoman2))

    inde()
def dao_excel():
    is_dao = input('创建文档class2.xlsx在file文件下，但别打开文档，好了写y')
    if is_dao == 'y':
        for row in range(1,people_count+1):
            if people.cell(row,1).value != None:
                da = load_workbook('./file/class2.xlsx')
                dao = da['Sheet1']
                dao.cell(row, 1).value = people.cell(row, 1).value
                dao.cell(row, 2).value = people.cell(row, 2).value
                dao.cell(row, 3).value = people.cell(row, 3).value
                dao.cell(row, 4).value = people.cell(row, 4).value
                da.save('./file/class2.xlsx')
                print('导出成功')
def inde():
    choose = eval(input('请输入你的学泽\n1.添加人员\n2.删除人员\n3.查询人员\n4.查所有人员\n5.导出所有人员成excel\n6.退出'))
    if choose == 1:
        add_person()
        inde()
    elif choose == 2:
        reversed()
        inde()
    elif choose == 3:
        cha()
    elif choose == 4:
        all_cha()
    elif choose == 5:
        dao_excel()
    elif choose == 6:
        exi()


print('欢迎使用汪星人人员记录系统,版本号0.0.3')
choose_account = eval(input('登录\n1.注册\n2.我已有汪星人账号，我要登录'))
if choose_account == 1:
    want_account = input('请输入你希望的用户名')
    account_pwd.cell(account_account + 1, 1).value = want_account
    want_pwd = input('请输入你希望的密码')
    account_pwd.cell(account_account + 1, 2).value = want_pwd
    middle = want_account + want_pwd
    wb.create_sheet(middle)
    people = wb[middle]
    people.cell(1, 1).value = '人员'
    people.cell(1, 2).value = '手机号'
    people.cell(1, 3).value = '邮箱'
    people.cell(1, 4).value = '性别'
    wb.save('./file/xlsx/class.xlsx')
    while True:
        people_count = people.max_row
        inde()
    del want_account
    del want_pwd
elif choose_account == 2:
    my_account = input('请输入你的用户名')
    my_pwd = input('请输入你的密码')
    middle2 = my_account + my_pwd
    people = wb[middle2]
    wb.save('./file/xlsx/class.xlsx')
    with open("savepwd.txt","w+") as pwdfile:
        pwdfile.write(my_account)
        pwdfile.write(my_pwd)
        
    while True:
        people_count = people.max_row
        inde()
# 尚未开发完毕
# elif choose_account == 3:
#     with open("savepwd.txt","r+") as pwdfile:
#         txtsheet = pwdfile.read()
#         people = wb[txtsheet]

wb.save('./file/xlsx/class.xlsx')
